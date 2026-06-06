"""B5: /jobs/{id}/export/{fmt} 匯出端點整合測試（TDD 先行）。"""
from __future__ import annotations

import csv
import io
import json

from fastapi.testclient import TestClient

from domain.entities import Job, Task
from infrastructure.container import Container
from infrastructure.web import create_app
from tests.fakes import FakeBrowserAgent, InMemoryTaskRepo


def _make_client(result: str | None = None):
    repo = InMemoryTaskRepo()
    task = Task.create(instruction="抓資料")
    job = Job.create(task_id=task.id)
    job.start_planning()
    job.start_running()
    if result is not None:
        job.succeed(result=result)
    repo.add_task(task)
    repo.add_job(job)
    container = Container.for_testing(repo=repo, agent=FakeBrowserAgent())
    app = create_app(container)
    return TestClient(app), job.id


def test_export_txt():
    client, job_id = _make_client(result="找到 3 筆資料")
    r = client.get(f"/jobs/{job_id}/export/txt")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/plain")
    assert "找到 3 筆資料" in r.text


def test_export_json_with_json_result():
    data = [{"name": "Alice", "score": 90}, {"name": "Bob", "score": 85}]
    client, job_id = _make_client(result=json.dumps(data))
    r = client.get(f"/jobs/{job_id}/export/json")
    assert r.status_code == 200
    parsed = r.json()
    assert parsed == data


def test_export_json_with_text_result_wraps():
    client, job_id = _make_client(result="純文字結果")
    r = client.get(f"/jobs/{job_id}/export/json")
    assert r.status_code == 200
    parsed = r.json()
    assert "result" in parsed or "output" in parsed or "純文字結果" in str(parsed)


def test_export_csv_with_list_of_dicts():
    data = [{"name": "Alice", "score": "90"}, {"name": "Bob", "score": "85"}]
    client, job_id = _make_client(result=json.dumps(data))
    r = client.get(f"/jobs/{job_id}/export/csv")
    assert r.status_code == 200
    reader = csv.DictReader(io.StringIO(r.text))
    rows = list(reader)
    assert len(rows) == 2
    assert rows[0]["name"] == "Alice"


def test_export_csv_with_text_fallback():
    client, job_id = _make_client(result="純文字")
    r = client.get(f"/jobs/{job_id}/export/csv")
    assert r.status_code == 200  # 純文字 fallback


def test_export_xlsx_with_list_of_dicts():
    data = [{"product": "A", "qty": 5}, {"product": "B", "qty": 3}]
    client, job_id = _make_client(result=json.dumps(data))
    r = client.get(f"/jobs/{job_id}/export/xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    # Verify it's a valid Excel file
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "product" in headers


def test_export_xlsx_non_list_returns_422():
    client, job_id = _make_client(result="純文字")
    r = client.get(f"/jobs/{job_id}/export/xlsx")
    assert r.status_code == 422


def test_export_unknown_format_returns_400():
    client, job_id = _make_client(result="ok")
    r = client.get(f"/jobs/{job_id}/export/pdf")
    assert r.status_code == 400


def test_export_no_result_returns_404():
    client, job_id = _make_client(result=None)  # job still running, no result
    r = client.get(f"/jobs/{job_id}/export/txt")
    assert r.status_code == 404
